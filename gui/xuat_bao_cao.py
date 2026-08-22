# -*- coding: utf-8 -*-
"""Xuất kết quả mô phỏng ra tệp báo cáo.

Chủ sở hữu: TV1.

Module nhận trạng thái hệ thống cùng kết quả kiểm tra an toàn rồi ghi ra tệp
để đưa vào quyển báo cáo. Hỗ trợ hai định dạng:

- **HTML** — không cần cài thêm thư viện nào. Mở tệp bằng trình duyệt rồi bấm
  In và chọn "Lưu thành PDF" là ra bản PDF tiếng Việt chuẩn dấu.
- **Excel** — cần ``pip install openpyxl``. Tiện cho TV8 khi lập bảng kết quả.

Vì sao không dùng reportlab để xuất thẳng PDF: các phông có sẵn của reportlab
không chứa dấu tiếng Việt, muốn đúng dấu phải nhúng phông ngoài khá lằng nhằng.
Đi đường HTML rồi để trình duyệt in ra PDF vừa gọn vừa không bao giờ lỗi phông.

Cách dùng từ giao diện của TV6::

    from gui.xuat_bao_cao import xuat_html
    xuat_html("ket-qua.html", trang_thai, ket_qua)
"""

from __future__ import annotations

import html
import os
from datetime import datetime
from typing import List, Optional

from engine.banker_types import BankerState, RequestResult, SafetyResult

TIEU_DE = "Bài tập lớn Hệ điều hành — Đề tài 9: Giải thuật Banker"


# --------------------------------------------------------------------------
# Dựng các mảnh HTML
# --------------------------------------------------------------------------

def _bang_ma_tran(trang_thai: BankerState) -> str:
    """Bảng Allocation / Max / Need gộp chung, giống bảng trong Chương 2."""
    m = trang_thai.m
    cot = "".join("<th>{}</th>".format(html.escape(t))
                  for t in trang_thai.ten_tai_nguyen)

    hang = []
    for i in range(trang_thai.n):
        o = '<td class="p">P{}</td>'.format(i)
        for nhom in (trang_thai.allocation[i], trang_thai.max[i], trang_thai.need[i]):
            o += "".join("<td>{}</td>".format(v) for v in nhom)
        hang.append("<tr>{}</tr>".format(o))

    return """
<table>
  <thead>
    <tr>
      <th rowspan="2">P</th>
      <th colspan="{m}">Allocation</th>
      <th colspan="{m}">Max</th>
      <th colspan="{m}">Need</th>
    </tr>
    <tr>{cot}{cot}{cot}</tr>
  </thead>
  <tbody>{hang}</tbody>
</table>""".format(m=m, cot=cot, hang="".join(hang))


def _bang_nhat_ky(ket_qua: SafetyResult) -> str:
    """Nhật ký từng vòng lặp của thủ tục kiểm tra an toàn."""
    if not ket_qua.nhat_ky:
        return "<p><i>Không có nhật ký.</i></p>"

    hang = []
    for b in ket_qua.nhat_ky:
        tt = "P{}".format(b.tien_trinh) if b.tien_trinh is not None else "—"
        hang.append(
            "<tr>"
            "<td>{}</td>".format(b.buoc)
            + "<td>{}</td>".format(b.work_truoc)
            + '<td class="p">{}</td>'.format(tt)
            + "<td>{}</td>".format(b.need if b.need is not None else "—")
            + "<td>{}</td>".format(b.allocation if b.allocation is not None else "—")
            + "<td>{}</td>".format(b.work_sau)
            + "</tr>"
        )

    return """
<table>
  <thead><tr>
    <th>Bước</th><th>Work trước</th><th>Tiến trình chọn</th>
    <th>Need</th><th>Allocation</th><th>Work sau</th>
  </tr></thead>
  <tbody>{}</tbody>
</table>""".format("".join(hang))


def _bang_yeu_cau(lich_su: List[RequestResult]) -> str:
    """Lịch sử các yêu cầu tài nguyên đã gửi trong phiên làm việc."""
    if not lich_su:
        return ""
    hang = "".join(
        '<tr><td>{}</td><td class="p">{}</td><td>{}</td></tr>'.format(
            i + 1, r.ket_luan.value, html.escape(r.ly_do))
        for i, r in enumerate(lich_su)
    )
    return """
<h2>4. Lịch sử yêu cầu tài nguyên</h2>
<table>
  <thead><tr><th>#</th><th>Kết luận</th><th>Lý do</th></tr></thead>
  <tbody>{}</tbody>
</table>""".format(hang)


CSS = """
body { font-family: "Times New Roman", serif; font-size: 13pt; line-height: 1.5;
       margin: 2cm 2cm 2cm 3cm; color: #000; }
h1 { font-size: 16pt; text-align: center; }
h2 { font-size: 14pt; margin-top: 22px; }
table { border-collapse: collapse; margin: 10px 0 16px; }
th, td { border: 1px solid #000; padding: 4px 10px; text-align: center;
         font-size: 12pt; }
th { background: #e8e8e8; }
td.p { font-weight: bold; text-align: left; }
.ket-luan { font-size: 14pt; font-weight: bold; padding: 8px 14px;
            border: 2px solid #000; display: inline-block; margin: 6px 0; }
.an-toan { background: #d9ead3; }
.khong-an-toan { background: #f4cccc; }
.chan-trang { color: #555; font-size: 11pt; margin-top: 26px;
              border-top: 1px solid #999; padding-top: 8px; }
@media print { body { margin: 0; } }
"""


# --------------------------------------------------------------------------
# Hàm công khai
# --------------------------------------------------------------------------

def xuat_html(duong_dan: str,
              trang_thai: BankerState,
              ket_qua: Optional[SafetyResult] = None,
              lich_su_yeu_cau: Optional[List[RequestResult]] = None) -> str:
    """Ghi báo cáo ra tệp HTML, trả về đường dẫn tuyệt đối.

    Mở tệp bằng trình duyệt rồi bấm In, chọn "Lưu thành PDF" để có bản PDF.
    """
    lich_su_yeu_cau = lich_su_yeu_cau or []

    if ket_qua is None:
        khoi_ket_luan = "<p><i>Chưa chạy kiểm tra an toàn.</i></p>"
        khoi_nhat_ky = ""
    else:
        if ket_qua.an_toan:
            khoi_ket_luan = (
                '<div class="ket-luan an-toan">AN TOÀN</div>'
                "<p>Chuỗi an toàn: <b>{}</b></p>".format(
                    html.escape(ket_qua.chuoi_dep()))
            )
        else:
            treo = ", ".join("P{}".format(i) for i in ket_qua.treo) or "—"
            khoi_ket_luan = (
                '<div class="ket-luan khong-an-toan">KHÔNG AN TOÀN</div>'
                "<p>Tiến trình còn treo: <b>{}</b></p>".format(treo)
            )
        khoi_nhat_ky = "<h2>3. Nhật ký từng bước</h2>" + _bang_nhat_ky(ket_qua)

    noi_dung = """<!doctype html>
<html lang="vi"><head><meta charset="utf-8">
<title>{tieu_de}</title><style>{css}</style></head>
<body>
<h1>{tieu_de}</h1>

<h2>1. Dữ liệu đầu vào</h2>
<p>Số tiến trình: <b>{n}</b> &nbsp;&middot;&nbsp;
   Số loại tài nguyên: <b>{m}</b></p>
<p>Available = <b>{available}</b> &nbsp;&middot;&nbsp;
   Total = <b>{total}</b></p>
{bang_ma_tran}

<h2>2. Kết luận</h2>
{ket_luan}

{nhat_ky}
{yeu_cau}

<p class="chan-trang">Kết xuất tự động lúc {thoi_gian}.</p>
</body></html>""".format(
        tieu_de=html.escape(TIEU_DE),
        css=CSS,
        n=trang_thai.n,
        m=trang_thai.m,
        available=trang_thai.available,
        total=trang_thai.total,
        bang_ma_tran=_bang_ma_tran(trang_thai),
        ket_luan=khoi_ket_luan,
        nhat_ky=khoi_nhat_ky,
        yeu_cau=_bang_yeu_cau(lich_su_yeu_cau),
        thoi_gian=datetime.now().strftime("%H:%M ngày %d/%m/%Y"),
    )

    with open(duong_dan, "w", encoding="utf-8") as f:
        f.write(noi_dung)
    return os.path.abspath(duong_dan)


def xuat_excel(duong_dan: str,
               trang_thai: BankerState,
               ket_qua: Optional[SafetyResult] = None) -> str:
    """Ghi báo cáo ra tệp Excel. Cần ``pip install openpyxl``."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        raise RuntimeError("Chưa cài openpyxl. Chạy lệnh: pip install openpyxl")

    m = trang_thai.m
    wb = Workbook()
    ws = wb.active
    ws.title = "Ma tran"

    ws["A1"] = TIEU_DE
    ws["A1"].font = Font(bold=True, size=13)

    ws["A3"] = "Available"
    for j, v in enumerate(trang_thai.available):
        ws.cell(row=3, column=2 + j, value=v)

    ws["A5"] = "P"
    for k, nhom in enumerate(("Allocation", "Max", "Need")):
        o = ws.cell(row=4, column=2 + k * m, value=nhom)
        o.alignment = Alignment(horizontal="center")
        o.font = Font(bold=True)
        for j, t in enumerate(trang_thai.ten_tai_nguyen):
            ws.cell(row=5, column=2 + k * m + j, value=t).font = Font(bold=True)

    for i in range(trang_thai.n):
        r = 6 + i
        ws.cell(row=r, column=1, value="P{}".format(i)).font = Font(bold=True)
        for k, nhom in enumerate((trang_thai.allocation[i],
                                  trang_thai.max[i],
                                  trang_thai.need[i])):
            for j, v in enumerate(nhom):
                ws.cell(row=r, column=2 + k * m + j, value=v)

    if ket_qua is not None:
        ws2 = wb.create_sheet("Nhat ky")
        ws2.append(["Bước", "Work trước", "Tiến trình", "Need",
                    "Allocation", "Work sau"])
        for c in ws2[1]:
            c.font = Font(bold=True)
        for b in ket_qua.nhat_ky:
            ws2.append([
                b.buoc,
                str(b.work_truoc),
                "P{}".format(b.tien_trinh) if b.tien_trinh is not None else "-",
                str(b.need),
                str(b.allocation),
                str(b.work_sau),
            ])
        ws2.append([])
        ws2.append(["Kết luận",
                    "AN TOÀN" if ket_qua.an_toan else "KHÔNG AN TOÀN"])
        if ket_qua.an_toan:
            ws2.append(["Chuỗi an toàn", ket_qua.chuoi_dep()])

    wb.save(duong_dan)
    return os.path.abspath(duong_dan)
